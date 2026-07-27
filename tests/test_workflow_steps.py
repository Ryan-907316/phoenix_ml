"""End-to-end test of the WorkflowSession/step-function layer — the path the
GUI actually drives (workflow.py's standalone entry point is a separate,
simpler path already covered by test_workflow_smoke.py).

This layer is where several real bugs have lived (the interpretability
visual-restriction wiring, per-target monotonicity gating), because it's the
glue between every module — and until this file existed it had zero automated
coverage despite being the daily-use path.

Config is trimmed for speed (one fast model, random-search HPO only with 4
iterations, conformal-only UQ, metrics-only interpretability) but the step
sequence is the real one: preprocessing -> training -> interpretability
(before) -> HPO -> UQ (after) -> report, with the report step also persisting
the deployable predictor exactly as a UI run would.
"""
import os

import matplotlib
matplotlib.use("Agg")

import pandas as pd
from openpyxl import load_workbook

from phoenix_ml.persistence import PhoenixPredictor
from phoenix_ml.workflow_steps import (
    WorkflowSession,
    run_step_hpo,
    run_step_interpretability_before,
    run_step_perl,
    run_step_preprocessing,
    run_step_report,
    run_step_training,
    run_step_uq_after,
    run_step_uq_before,
)


def test_reset_results_clears_every_result_but_keeps_settings():
    """Regression test for a real bug: running dataset A with PERL enabled, then
    dataset B without re-running every step, carried dataset A's stale PERL
    section (and other results) into dataset B's report. reset_results() must
    clear every per-run result while leaving user-configured settings alone."""
    session = WorkflowSession(
        dataset_path="a.csv", output_dir="out", targets=["T"],
        selected_models=["KNeighbors Regressor"], random_seed=7,
        monotonic_constraints={"T": {"f": 1}},
    )
    # Populate every result field as if a previous run had completed.
    session.preprocessing_results = {"stub": True}
    session.training_results = {"stub": True}
    session.uq_before = ({"stub": True}, {})
    session.uq_after = ({"stub": True}, {})
    session.interpretability_before = ({"stub": True}, {})
    session.interpretability_after = ({"stub": True}, {})
    session.hpo_results = {"stub": True}
    session.cv_results = {"stub": True}
    session.perl_results = {"stub": True}
    session.perl_config = {"stub": True}
    session.perl_output_df = pd.DataFrame({"a": [1]})
    session.cleaning_summary = {"export_path": "x"}
    orig_cleaning_summary = session.cleaning_summary
    session.metrics["random"] = {"stub": True}
    session.params["random"] = {"stub": True}
    session.total_elapsed = 123.4
    session.step_timings = [("Preprocessing", 1.0)]
    session.images_dir = "/old/images"
    session.report_dir = "/old/report"
    session.xlsx_path = "/old/x.xlsx"
    session.pdf_path = "/old/r.pdf"
    session.models_dir = "/old/models"

    session.reset_results()

    for field in ["preprocessing_results", "training_results", "uq_before", "uq_after",
                  "interpretability_before", "interpretability_after", "hpo_results",
                  "cv_results", "perl_results", "perl_config", "perl_output_df",
                  "images_dir", "report_dir", "xlsx_path",
                  "pdf_path", "models_dir"]:
        assert getattr(session, field) is None, f"{field} was not cleared"
    assert session.metrics == {"default": {}, "random": {}, "hyperopt": {}, "skopt": {}}
    assert session.params == {"default": {}, "random": {}, "hyperopt": {}, "skopt": {}}
    assert session.total_elapsed == 0.0
    assert session.step_timings == []

    # Settings must survive untouched.
    assert session.dataset_path == "a.csv"
    assert session.output_dir == "out"
    assert session.targets == ["T"]
    assert session.selected_models == ["KNeighbors Regressor"]
    assert session.random_seed == 7
    assert session.monotonic_constraints == {"T": {"f": 1}}
    # cleaning_summary is NOT a per-run pipeline result -- it's set by a separate,
    # out-of-band UI action (the Clean tab export) that always happens BEFORE the
    # user runs Preprocessing on the exported file. Clearing it here made the
    # cleaning-removed-rows report/model-card content permanently unreachable in
    # normal use (regression test: test_cleaning_summary_survives_a_preprocessing_rerun_on_the_same_dataset).
    assert session.cleaning_summary is orig_cleaning_summary


def test_rerunning_preprocessing_clears_stale_results_from_a_previous_dataset(
    tmp_path, synthetic_dataset_csv,
):
    """End-to-end version of the same bug: a session that already has PERL/HPO
    results from a first dataset must have them gone the moment preprocessing
    runs again for a second (different) dataset — without the user needing to
    click anything else. This is the automatic, can't-forget-it fix; the manual
    Reset button in the UI calls the same reset_results() method directly."""
    session = WorkflowSession(
        dataset_path=str(synthetic_dataset_csv), output_dir=str(tmp_path),
        targets=["Target"], selected_models=["KNeighbors Regressor"], random_seed=0,
        show_target_vs_target=False, show_features_vs_targets=False,
        show_boxplots=False, show_distance_corr=False,
        show_multicollinearity=False, plot_pca_enabled=False, feat_sel_enabled=False,
    )
    # Simulate a completed prior run on a DIFFERENT dataset.
    session.perl_results = {"stale": "from a previous dataset"}
    session.perl_config = {"stale": True}
    session.hpo_results = {"stale": True}
    session.total_elapsed = 999.0
    session.step_timings = [("Hyperparameter Optimisation", 999.0)]
    old_images_dir = str(tmp_path / "stale_images")
    session.images_dir = old_images_dir

    run_step_preprocessing(session)

    assert session.perl_results is None
    assert session.perl_config is None
    assert session.hpo_results is None
    assert session.total_elapsed == 0.0
    assert session.step_timings == []
    # The stale path must have been recomputed, not silently kept.
    assert session.images_dir != old_images_dir
    assert session.preprocessing_results is not None


def test_selected_warns_once_for_a_model_name_absent_from_all_models(capsys):
    """Regression test for a real risk: session.selected silently dropped any
    selected_models entry not present in ALL_MODELS (e.g. a stale saved
    config referencing a model since renamed/removed) — the pipeline just
    ran with fewer models than configured, with nothing telling the user."""
    session = WorkflowSession(
        dataset_path="a.csv", output_dir="out", targets=["T"],
        selected_models=["KNeighbors Regressor", "Not A Real Model"],
    )
    models = session.selected
    assert list(models.keys()) == ["KNeighbors Regressor"]
    out = capsys.readouterr().out
    assert "[WARN]" in out and "Not A Real Model" in out

    # Repeated access must not reprint the same warning.
    _ = session.selected
    _ = session.selected
    assert capsys.readouterr().out == ""


def test_perl_warns_when_every_target_is_skipped(tmp_path, synthetic_dataset_csv, capsys):
    """Regression test for a real risk: a PERL run where every target hits a
    per-target skip (e.g. a reconstruction_map typo pointing at a physics
    column that was never produced) left perl_results = {}, indistinguishable
    from "PERL was never attempted" — the report section silently vanishes
    (gated on `if session.perl_results:`, falsy for {}) with nothing telling
    the user PERL ran and found nothing usable."""
    from phoenix_ml.physics_expressions import save_physics_config

    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    run_step_preprocessing(session)
    run_step_training(session)

    config_path = str(tmp_path / "physics_config.json")
    save_physics_config(
        config_path, expressions=[], output_cols_text="",
        reconstruction_map={"Target": "nonexistent_physics_col"},
    )
    session.perl_config_path = config_path

    run_step_perl(session)

    assert session.perl_results == {}
    assert "every target was skipped" in capsys.readouterr().out


def _tiny_session(dataset_csv, output_dir):
    return WorkflowSession(
        dataset_path=str(dataset_csv),
        output_dir=str(output_dir),
        targets=["Target"],
        selected_models=["KNeighbors Regressor"],
        random_seed=0,
        # Keep preprocessing plot generation off — the report must cope with
        # every optional figure absent (the UI exposes these as checkboxes).
        show_target_vs_target=False, show_features_vs_targets=False,
        show_boxplots=False, show_distance_corr=False,
        show_multicollinearity=False, plot_pca_enabled=False,
        feat_sel_enabled=False,
        # Metrics-only interpretability: no ICE/PDP/ALE/SHAP visuals, just a
        # tiny Morris pass feeding the comparable-metrics table.
        interpretability_settings=dict(
            test_sample_size=40, background_sample_size=5,
            subsample=10, grid_resolution=5,
            show_ice_pdp=False, show_ale=False, show_shap_summary=False,
            show_shap_dependence=False, show_shap_waterfall=False,
            show_sensitivity_morris=True, sensitivity_morris_trajectories=4,
            sensitivity_morris_levels=4, show_sensitivity_sobol=False,
        ),
        methods_to_run=["random"],
        sampling_method="Random",
        n_iter=4,
        n_jobs=1,
        early_stopping=None,
        uq_settings=dict(
            uq_method="Conformal", n_bootstrap=2, confidence_interval=95,
            calibration_frac=0.1, subsample_test_size=16, n_jobs=1,
            include_gp_posterior=False, calibration_enabled=True,
        ),
    )


def test_session_pipeline_end_to_end(tmp_path, synthetic_dataset_csv):
    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")

    # Prerequisite gates must open in order, exactly as the UI's step
    # checkboxes rely on them doing.
    assert session.can_run_preprocessing()
    assert not session.can_run_training()

    run_step_preprocessing(session)
    assert session.preprocessing_results is not None
    assert session.can_run_training()
    assert not session.can_run_hpo()

    run_step_training(session)
    assert session.training_results is not None
    assert not session.training_results["results_df"].empty
    assert session.can_run_hpo()
    assert not session.can_run_uq_after()          # needs HPO first

    run_step_interpretability_before(session)
    metrics_df, figures = session.interpretability_before
    # One comparable-metrics row per (model, target).
    assert len(metrics_df) == 1
    assert metrics_df.iloc[0]["Model"] == "KNeighbors Regressor"

    run_step_hpo(session)
    assert session.hpo_results is not None
    best = session.hpo_results["best_models_per_target"]
    assert "Target" in best
    assert session.can_run_uq_after()

    run_step_uq_after(session)
    uq_df, uq_figs = session.uq_after
    assert not uq_df.empty
    assert (uq_df["UQ Method"] == "Conformal").all()

    assert session.can_generate_report()
    run_step_report(session)

    # The two artifacts every UI run hands the user: the PDF report and the
    # deployable predictor.
    assert os.path.isfile(session.pdf_path)
    assert os.path.getsize(session.pdf_path) > 10_000   # a real multi-page PDF

    # Regression test: sheets used to be written with openpyxl's unset default
    # column width, clipping headers/values when the workbook was opened.
    # Every column must now be sized to fit its widest cell.
    assert session.xlsx_path and os.path.isfile(session.xlsx_path)
    wb = load_workbook(session.xlsx_path)
    assert wb.sheetnames, "workbook must contain at least one sheet"
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            letter = col_cells[0].column_letter
            width = sheet.column_dimensions[letter].width
            longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
            assert width is not None, f"{sheet.title}!{letter} has no explicit column width"
            if longest:
                assert width >= min(longest, 60), (
                    f"{sheet.title}!{letter} width {width} too narrow for its longest value ({longest} chars)"
                )

    predictor_files = [f for f in os.listdir(session.models_dir)
                       if f == "phoenix_ml Predictor.pkl"]
    assert len(predictor_files) == 1

    # And that predictor must actually work on fresh raw-feature data.
    predictor = PhoenixPredictor.load(os.path.join(session.models_dir, predictor_files[0]))
    raw = pd.read_csv(synthetic_dataset_csv).drop(columns=["Target"])
    preds = predictor.predict(raw)
    assert list(preds.columns) == ["Target"]
    assert len(preds) == len(raw)
    assert preds["Target"].notna().all()


def test_report_does_not_fabricate_settings_for_steps_that_never_ran(tmp_path, synthetic_dataset_csv):
    """Regression test for a real bug found by the user on a genuine run: with
    HPO's checkbox off, session.methods_to_run stayed at its untouched dataclass
    default (all three methods) since _sync_session only refreshes it when the
    HPO step is enabled -- and both the model card and the Reproducibility
    Metadata JSON trusted that stale value unconditionally, fabricating "HPO used
    Random Search/Hyperopt/Scikit-Optimize" when HPO never ran at all. Same class
    of bug existed for uq_settings/interpretability_settings in the metadata JSON.
    Only preprocessing + training + report run here -- HPO, UQ, Interpretability,
    and CV are all left off, mirroring the exact scenario that surfaced this."""
    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    run_step_preprocessing(session)
    run_step_training(session)
    assert session.can_generate_report()
    run_step_report(session)

    import json
    meta_path = [f for f in os.listdir(session.models_dir) if f.endswith("Metadata.json")][0]
    with open(os.path.join(session.models_dir, meta_path)) as f:
        meta = json.load(f)
    assert meta["settings"]["hpo"] is None
    assert meta["settings"]["uq"] is None
    assert meta["settings"]["interpretability"] is None

    # The model card itself must still generate (just correctly, not fabricated --
    # the exact "Default" label is covered directly by _hpo_label_for_row's own
    # tests in test_model_cards.py; reportlab's PDF text isn't stored as plain
    # matchable substrings, so asserting on rendered bytes here isn't reliable).
    card_path = os.path.join(session.models_dir, "Model Cards", "phoenix_ml Model Card_Target.pdf")
    assert os.path.isfile(card_path)
    assert session.hpo_results is None  # confirms this scenario genuinely didn't run HPO


def test_excel_export_sheet_names_reflect_default_hyperparameters_not_before_hpo(
    tmp_path, synthetic_dataset_csv,
):
    """User-reported: the PDF report's "Before HPO" -> "Default Hyperparameters"
    relabelling didn't extend to the Excel export, which still hardcoded "UQ Before
    HPO"/"Interpretability Before HPO" sheet names (and Summary sheet "Contents"
    text) regardless of whether HPO actually ran. Excel sheet names are capped at
    31 characters, so the full "Default Hyperparameters" label can't be used
    verbatim in "Interpretability " + label (41 chars) -- must use a short,
    length-safe form for the sheet name while the Summary sheet's free-text
    "Contents" column can carry the full label."""
    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    session.hpo_enabled = False
    run_step_preprocessing(session)
    run_step_training(session)
    run_step_uq_before(session)
    assert session.can_generate_report()
    run_step_report(session)

    assert session.xlsx_path and os.path.isfile(session.xlsx_path)
    wb = load_workbook(session.xlsx_path)
    assert "UQ Default HP" in wb.sheetnames
    assert "UQ Before HPO" not in wb.sheetnames
    for name in wb.sheetnames:
        assert len(name) <= 31

    summary_ws = wb["Summary"]
    header = [c.value for c in summary_ws[1]]
    sheet_col = header.index("Sheet") + 1
    contents_col = header.index("Contents") + 1
    rows = {row[sheet_col - 1].value: row[contents_col - 1].value
            for row in summary_ws.iter_rows(min_row=2)}
    assert "Default Hyperparameters" in rows["UQ Default HP"]


def test_report_relabels_before_hpo_as_default_hyperparameters_when_hpo_never_ran(
    tmp_path, synthetic_dataset_csv,
):
    """User-requested (2026-07-22): "Before HPO" is only an accurate label when
    there's an "after" to contrast it against. This test's session leaves
    session.hpo_enabled at its default (True, i.e. the HPO checkbox is still
    ticked) but never actually calls run_step_hpo -- exactly the case where
    run_step_uq_before() (which runs before this session's real HPO outcome is
    known) still bakes in "Before HPO" at the source, so the correction has to
    happen at report-assembly time instead (see the sibling test below for the
    newer, more common source-level case where hpo_enabled is False from the
    start). Verified via a spy on handle_uq_reporting_section rather than
    parsing rendered PDF bytes, per this file's established technique."""
    from unittest.mock import patch
    from phoenix_ml import workflow_steps as ws

    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    assert session.hpo_enabled is True  # this test's scenario: ticked, but never run
    run_step_preprocessing(session)
    run_step_training(session)
    run_step_uq_before(session)
    assert session.uq_before is not None
    # The raw session data is unaffected by the report-time fix -- still literally
    # "Before HPO" at the source, since run_step_uq_before() has no way to know at
    # this point that HPO won't actually run later this session.
    assert set(session.uq_before[0]["Stage"].unique()) == {"Before HPO"}
    assert any(" - Before HPO" in k for k in session.uq_before[1])

    with patch.object(
        ws, "handle_uq_reporting_section", wraps=ws.handle_uq_reporting_section,
    ) as spy:
        run_step_report(session)
        stage_arg = spy.call_args.args[2]
        assert stage_arg == "Default Hyperparameters"
        uq_df_arg, uq_figs_arg = spy.call_args.args[0], spy.call_args.args[1]
        assert set(uq_df_arg["Stage"].unique()) == {"Default Hyperparameters"}
        # Figure captions (their dict keys) must be corrected too, not just the
        # Stage column -- these are rendered as plain caption text above each
        # model's UQ figure and would otherwise contradict the section heading.
        assert all(" - Before HPO" not in k for k in uq_figs_arg)
        assert any(" - Default Hyperparameters" in k for k in uq_figs_arg)
    assert session.hpo_results is None  # confirms this scenario genuinely didn't run HPO


def test_uq_before_labelled_correctly_at_the_source_when_hpo_disabled_upfront(
    tmp_path, synthetic_dataset_csv,
):
    """Common case (unlike the test above): the HPO checkbox is off from the
    start (session.hpo_enabled=False, as _sync_session would set it), so
    run_step_uq_before() itself -- not a report-time patch -- gets the Stage
    column and figure captions right immediately."""
    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    session.hpo_enabled = False
    run_step_preprocessing(session)
    run_step_training(session)
    run_step_uq_before(session)
    assert set(session.uq_before[0]["Stage"].unique()) == {"Default Hyperparameters"}
    assert any(" - Default Hyperparameters" in k for k in session.uq_before[1])
    assert all(" - Before HPO" not in k for k in session.uq_before[1])


def test_report_keeps_before_hpo_label_when_hpo_did_run(tmp_path, synthetic_dataset_csv):
    """Counterpart to the test above: with HPO enabled and producing real
    results, the "Before HPO" label is accurate and must be left alone."""
    from unittest.mock import patch
    from phoenix_ml import workflow_steps as ws

    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    run_step_preprocessing(session)
    run_step_training(session)
    run_step_uq_before(session)
    run_step_hpo(session)
    assert session.hpo_results is not None

    with patch.object(
        ws, "handle_uq_reporting_section", wraps=ws.handle_uq_reporting_section,
    ) as spy:
        run_step_report(session)
        assert spy.call_args.args[2] == "Before HPO"


def test_report_relabels_default_hyperparameters_as_before_hpo_when_hpo_force_run_after(
    tmp_path, synthetic_dataset_csv,
):
    """Regression test for the reverse direction of the asymmetric fix above: the
    UI's per-row "Run" button can force a step to run "regardless of that row's
    Enable checkbox" (forced_step mechanism). A user can run UQ-Before with the HPO
    checkbox off (baking "Default Hyperparameters" into the Stage column/figure
    captions at the source), then separately force-run HPO anyway. The previous
    correction only handled label=="Before HPO" baked but hr empty; this is the
    opposite -- label=="Default Hyperparameters" baked but hr now non-empty -- and
    was previously left uncorrected, contradicting the Executive Summary's
    fact-based "before hyperparameter optimisation" prose and the "Before HPO"
    section heading on the same page."""
    from unittest.mock import patch
    from phoenix_ml import workflow_steps as ws

    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    session.hpo_enabled = False  # simulates the checkbox being off when UQ-Before ran
    run_step_preprocessing(session)
    run_step_training(session)
    run_step_uq_before(session)
    assert set(session.uq_before[0]["Stage"].unique()) == {"Default Hyperparameters"}

    session.hpo_enabled = True  # simulates the user then force-running HPO anyway
    run_step_hpo(session)
    assert session.hpo_results is not None

    with patch.object(
        ws, "handle_uq_reporting_section", wraps=ws.handle_uq_reporting_section,
    ) as spy:
        run_step_report(session)
        assert spy.call_args.args[2] == "Before HPO"
        uq_df_arg, uq_figs_arg = spy.call_args.args[0], spy.call_args.args[1]
        assert set(uq_df_arg["Stage"].unique()) == {"Before HPO"}
        assert all(" - Default Hyperparameters" not in k for k in uq_figs_arg)
        assert any(" - Before HPO" in k for k in uq_figs_arg)


def test_cleaning_summary_survives_a_preprocessing_rerun_on_the_same_dataset(
    tmp_path, synthetic_dataset_csv,
):
    """Regression test for a real bug found via edge-case sweep testing: the only
    sensible order of operations is clean -> export a cleaned CSV -> point
    Preprocessing at that exported file -> run Preprocessing -> ... -> Report, since
    Preprocessing must consume the already-cleaned data. But reset_results() (called
    unconditionally at the start of run_step_preprocessing) used to null out
    session.cleaning_summary every time, which always runs AFTER the Clean tab sets
    it and BEFORE Report Generation ever reads it -- so the "rows removed during
    cleaning" content could never fire in real usage, no matter how much cleaning
    actually removed. Staleness across a genuine dataset switch is already handled
    separately, by comparing session.dataset_path against the summary's own
    export_path at report time (see run_step_report) -- that check is suffient on
    its own and doesn't need reset_results() as a backstop."""
    from unittest.mock import patch

    session = _tiny_session(synthetic_dataset_csv, tmp_path / "out")
    # Simulate the Clean tab's Export button: cleaning_summary set BEFORE
    # Preprocessing runs on the exported (== this test's dataset) file, exactly as
    # a real user session would sequence it.
    session.cleaning_summary = {
        "rows_before": 100, "rows_after": 80,
        "actions": ["[FIX ] Outlier rows removed (IQR): 20"],
        "export_path": os.path.abspath(str(synthetic_dataset_csv)),
    }

    run_step_preprocessing(session)
    # The bug: this used to be None here, unconditionally.
    assert session.cleaning_summary is not None
    assert session.cleaning_summary["rows_before"] == 100

    run_step_training(session)
    with patch(
        "phoenix_ml.workflow_steps.generate_model_cards", wraps=None,
    ) as mock_cards:
        from phoenix_ml.model_cards import generate_model_cards as real_generate_model_cards
        mock_cards.side_effect = real_generate_model_cards
        run_step_report(session)
        _, kwargs = mock_cards.call_args
        assert kwargs["cleaning_summary"] == session.cleaning_summary
